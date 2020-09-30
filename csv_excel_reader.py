import csv
import pandas as pd


def get_list_from_csv(filename, column_index, ignore_header=True):
    tags = []
    with open(filename, 'r') as csv_file:
        reader = csv.reader(csv_file)
        for i, row in enumerate(reader):
            tags.append(row[column_index])
    return tags

def get_list_from_excel(filename, column_index, ignore_header= True):
    from openpyxl import load_workbook
    tags = []
    wb=load_workbook(filename = 'so_tags.xlsx')
    sheet=wb.worksheets[0]  # Loading first sheet.
    for row_num, row in enumerate(sheet.iter_rows()):
        if row_num == 0:
            continue  # skipping header
        # each row is a collection of cell objects. 
        # Nested loop is needed
        for col_num, cell in enumerate(row):
            if col_num == 1:
                tags.append(cell.value)
    return tags



def get_list_from_csv_pd(filename, column_name=None, column_index=None, ignore_header=True):
    df = pd.read_csv('so_tags.csv')
    if column_name:
        return df[column_name].values.tolist()
    elif column_index:
        return df.iloc[:,column_index].values.tolist()
    else:
        raise Exception("Either column_name or column_index is required")


def get_list_from_excel_pd(filename, column_name=None, column_index=None, ignore_header=True):
    df = pd.read_excel('so_tags.xlsx')
    if column_name:
        return df[column_name].values.tolist()
    elif column_index:
        return df.iloc[:,column_index].values.tolist()
    else:
        raise Exception("Either column_name or column_index is required")

